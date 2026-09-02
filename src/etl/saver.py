import streamlit as st
import pandas as pd
import re
from openpyxl.utils import get_column_letter
from pathlib import Path


_SHEET_NAME_MAP: dict[str, str] = {
    "summary of sales by customer by item": "summary of sales",
    "purchase master report for all branches": "purchase master report",
    "discount by category by department": "disc by category by dep",
    "discount by description by employee": "discount by description",
    "inventory / summary of sales by customer by items": "inventory__summary of sales",
    "sales / summary of sales by customer by items": "sales__summary of sales" ,
    "discount by invoice with details": "discount by invoice",
    "discount by invoice with percentage": "discount by invoice with %",
}


def save_cleaned_data(cleaned: dict[str, object], raw_folder: str | Path, result_name: str = 'Cleaned Data.xlsx') -> Path | None:
    
    folder_path = Path(raw_folder)
    if not folder_path.exists() or not folder_path.is_dir():
        raise NotADirectoryError(f"Folder not found or not a directory: {folder_path}")

    workbook_path = folder_path / result_name

    try:
        with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
            for name, value in cleaned.items():
                if not isinstance(value, pd.DataFrame):
                    continue

                sheet_name = _SHEET_NAME_MAP.get(name, name)[:31]

                value.to_excel(writer, sheet_name=sheet_name, index=False)

                ws = writer.book[sheet_name]

                # Filters on header row
                ws.auto_filter.ref = ws.dimensions

                # Freeze first row
                ws.freeze_panes = "A2"

                # Adjust column widths
                for col_idx, column in enumerate(value.columns, start=1):
                    max_length = len(str(column))

                    for cell in ws[get_column_letter(col_idx)]:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))

                    ws.column_dimensions[get_column_letter(col_idx)].width = min(
                        max_length + 6,
                        50,  # prevents absurdly wide columns
                    )

                if re.search(r" \d+$", sheet_name):
                    writer.book[sheet_name].sheet_state = "hidden"

    except PermissionError:
        st.write(f"❌❌ please close the excel file '{result_name}' and try again.")
        st.stop()
        return None

    return workbook_path